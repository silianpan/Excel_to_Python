import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-12-3-1.xlsx') #读取工作簿。
ws=wb['优秀员工表'] #读取工作表。
num=0 #初始化num变量为0。
if not '优秀员工表2' in wb.sheetnames: #如果'优秀员工表2'工作表不存在。
    nws=wb.create_sheet('优秀员工表2') #则新建'优秀员工表2'工作表。
else:#否则。
    wb.remove(wb['优秀员工表2']) #先删除'优秀员工表2'工作表。
    nws = wb.create_sheet('优秀员工表2') #然后再新建'优秀员工表2'工作表。
nws.append(['序号','部门','姓名']) #在新建工作表中写入表头。
for row in ws.iter_rows(min_row=2): #从第2行开始，将每行数据循环赋值给row变量。
    print(row,row[1].value.split('、'))
    for val in row[1].value.split('、'): #获取每行数据的第1个元素，以顿号为分隔符拆分成列表。
        num +=1 #对num变量累加1。
        nws.append([num,row[0].value,val]) #将num是序号，row[0]是部门，val是拆分出来的每个姓名，组成列表写入新工作表。
# wb.save('Chapter-12-3-1.xlsx') #保存工作簿。
