import openpyxl #导入库。
wb=openpyxl.load_workbook('demo.xlsx') #读取工作簿。
ws=wb['业绩表'] #读取工作表。
if not '汇总结果' in wb.sheetnames: #如果'汇总结果'工作表在工作簿中不存在。
    nws=wb.create_sheet('汇总结果') #则新建'汇总结果'工作表。
else: #否则如果存在。
    wb.remove(wb['汇总结果']) #先删除'汇总结果'工作表。
    nws = wb.create_sheet('汇总结果') #然后再新建'汇总结果'工作表。
nws.append(['部门','总业绩','最高业绩','最低业绩','计数']) #给新建的工作表添加表头。
dic={} #创建dic空字典
for row in list(ws.values)[1:]: #按行循环获取ws工作表中的每一行数据。
    if not row[0] in dic.keys(): #如果row[0]在dic字典中不存在。
        dic[row[0]]=list(row[2:]) #那么以row[0]为键，row[2:]为值，创建新的键值对。
    else: #否则如果存在。
        dic[row[0]].extend(row[2:]) #则将row[2:]添加到row[0]键对应的值中。
for key,item in dic.items(): #循环获取dic字典的所有键和值。
    nws.append([key,sum(item),max(item),min(item),len(item)]) #汇总item列表中的数据，写入新工作表。
wb.save('demo.xlsx') #保存工作簿。
