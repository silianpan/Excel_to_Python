import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-12-6-1.xlsx')#读取工作簿。
ws=wb['业绩表']#读取工作表。
if not '最高业绩表' in wb.sheetnames:#如果要新建的表在工作簿中不存在。
    nws=wb.create_sheet('最高业绩表')#则新建工作表。
else:#否则，也就是如果存在。
    wb.remove(wb['最高业绩表'])#则先删除此工作用。
    nws=wb.create_sheet('最高业绩表')#然后再新建工作表。
nws.append(['部门','姓名','月份','业绩'])#在新建的工作表写入表头。
for col in ws.iter_cols(min_row=2,min_col=3):#从C2单元格开始，按列获取每列的数据。
    lst=[v.value for v in col]#将每列的单元格对象使用列表推导式转换为值。
    max_val=max(lst)#在lst列表中求出最大值。
    for cell in filter(lambda x:x.value==max_val,col):#筛选出每列等于最大值的单元格，然后循环给cell变量。
       tup=ws[cell.row][:2]+(ws.cell(1,cell.column),cell)#获取最大值所在行的部门、姓名、月份、业绩值单元格。组合成一个元组。
       nws.append([v.value for v in tup])#对tup元组做列表推导式，获取单元格的值形成列表。再写入新建的表。
wb.save('Chapter-12-6-1.xlsx')#保存工作簿。
