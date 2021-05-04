import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-16-1.xlsx') #读取工作簿。
ws=wb.active #读取活动工作表。
print(ws.min_row) #最小行号。
print(ws.max_row) #最大行号。
print(ws['B4'].row) #指定单元格行号。
print(ws.values) #按行获取工作表全部单元格的值。
print(ws.rows) #按行获取工作表全部单元格对象。
print(ws.iter_rows(min_row=3,min_col=2)) #从3行、2列开始，按行获取工作表已用单元格对象。

