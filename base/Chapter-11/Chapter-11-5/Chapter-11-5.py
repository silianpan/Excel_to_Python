import openpyxl #导入库。
nwb=openpyxl.Workbook() #新建工作簿。
nwb.create_sheet() #在新建工作簿最后新建工作表。
nwb.create_sheet('工资表') #在新建工作簿最后新建'工资表'的工作表。
nwb.create_sheet('汇总表',0) #在新建工作簿中第0个工作表前面新建'汇总表'工作表。
nwb.save('Chapter-11-5-1.xlsx') #保存工作簿。
