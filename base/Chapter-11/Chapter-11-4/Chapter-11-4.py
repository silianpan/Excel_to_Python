import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-4-1.xlsx') #读取工作簿。
wb.create_sheet() #在工作簿最后新建工作表。
wb.create_sheet('7月') #在工作簿最后新建'7月'工作表。
wb.create_sheet('3月',2) #在工作簿的第2个工作表前新建工作表。
wb.save('Chapter-11-4-2.xlsx') #保存工作簿。

