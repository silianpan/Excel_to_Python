import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-17-1.xlsx') #读取工作簿。
ws=wb.active #读取活动工作表。
print(ws.min_column) #最小列号。
print(ws.max_column) #最大列号。
print(ws['B4'].column) #指定单元格列号。
print(ws.values) #按列获取工作表全部单元格的值。
print(ws.columns) #按列获取工作表全部单元格对象。
print(ws.iter_cols(min_row=3,min_col=2)) #从3行、2列开始，按列获取工作表已用单元格对象。
