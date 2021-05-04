import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-15-1.xlsx') #读取工作簿。
ws=wb.active #读取活动工作表。
print(ws['A1:B2'],ws['A1':'B2']) #获取选择区域的2种方法。
print(ws['1:2'],ws['1':'2'],ws[1:2]) #获取指定行区域的3种方法。
print(ws['A:B'],ws['A':'B']) #获取指定列区域的2种方法。

