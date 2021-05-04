import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-14-1.xlsx') #读取工作簿。
ws=wb.worksheets[0] #读取第0个工作表。
print(ws['A1'],ws['A1'].value) #获取A1单元格对象，及A1单元格对象的值。
print(ws.cell(1,1),ws.cell(1,1).value) #获取A1单元格对象，及A1单元格对象的值。

