import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-18-1.xlsx') #读取工作簿。
ws=wb.active #读取活动工作表
ws['A4']='小明';ws['B4']=100;ws['C4']=89#按A1方法写入单元格。
ws.cell(5,1,'小松');ws.cell(5,2,97);ws.cell(5,3,96) #按R1C1方法写入单元格。
wb.save('Chapter-11-18-2.xlsx') #另存工作簿。
