import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-21-1.xlsx') #读取工作簿。
ws=wb['成绩表'] #读取工作表。
score=ws.iter_rows(min_row=2,min_col=2,max_col=ws.max_column-1) #获取工作表指定区域。
for row in score: #循环每行分数单元格。
    total=sum([v.value for v in row]) #获取每行单元格总分。
    row[-1].offset(0,1).value=total #将总分写入E列单元格 方法1。
    ws['E'+str(row[1].row)]=total #将总分写入E列单元格 方法2。
    ws.cell(row[1].row,5,total) #将总分写入E列单元格 方法3。
wb.save('Chapter-11-21-1.xlsx') #保存工作簿。
