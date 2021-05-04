import openpyxl #导入库。
wb=openpyxl.load_workbook('Chapter-11-27-1.xlsx',data_only=True) #读取工作簿。
ws=wb['工资表'] #读取工作表。
for cell in ws['B'][2:]: #循环指列的单元格对象。
    ws.insert_rows(cell.row) #在指定行插入空白行。
    for col_num in range(1,7): #循环序列数1~6做为列号。
        tit=ws.cell(1,col_num).value #获取表头的值。
        ws.cell(cell.row-1,col_num,tit) #将表头的值写入写插入的空白行单元格。
wb.save('Chapter-11-27-2.xlsx') #另存工作簿。
